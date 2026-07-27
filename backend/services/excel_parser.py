"""商品资料文件解析模块

支持 .xlsx / .xls / .csv 格式，自动提取商品字段。
预期列名（兼容中英文）：商品名称、价格、分类、商品描述、规格参数、卖点
"""

import csv
from io import BytesIO, StringIO

from openpyxl import load_workbook

# 英文列名 → 中文列名 映射
COLUMN_MAP = {
    "product_name": ["商品名称", "产品名称", "品名", "name", "product_name", "title"],
    "price": ["价格", "售价", "定价", "price", "零售价"],
    "category": ["分类", "品类", "类目", "category", "cat"],
    "description": ["商品描述", "产品描述", "描述", "description", "desc"],
    "specs": ["规格参数", "规格", "参数", "specs", "规格型号", "specification"],
    "selling_points": [
        "卖点",
        "核心卖点",
        "产品卖点",
        "selling_points",
        "highlights",
        "亮点",
    ],
}


def _find_column(headers: list[str], field: str) -> int | None:
    """在表头行查找匹配的列索引"""
    candidates = COLUMN_MAP.get(field, [])
    for i, h in enumerate(headers):
        h_clean = str(h).strip().lower()
        if h_clean in [c.lower() for c in candidates]:
            return i
    return None


def _rows_to_products(rows: list[tuple], headers: list[str]) -> list[dict]:
    """将行数据转为产品字典列表"""
    col_index = {f: _find_column(headers, f) for f in COLUMN_MAP}
    products = []
    for row in rows:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        product = {}
        for field, idx in col_index.items():
            if idx is not None and idx < len(row):
                val = row[idx]
                product[field] = str(val).strip() if val is not None else ""
            else:
                product[field] = ""
        if product.get("product_name", ""):
            products.append(product)
    return products


def parse_excel(file_bytes: bytes) -> list[dict]:
    """解析 Excel 文件（.xlsx / .xls）"""
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c) if c is not None else "" for c in rows[0]]
    products = _rows_to_products(rows[1:], headers)
    wb.close()
    return products


def parse_csv(file_bytes: bytes) -> list[dict]:
    """解析 CSV 文件"""
    text = file_bytes.decode("utf-8-sig")
    reader = csv.reader(StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    headers = [str(c).strip() for c in rows[0]]
    products = _rows_to_products(rows[1:], headers)
    return products


def parse_product_file(file_bytes: bytes, filename: str) -> list[dict]:
    """根据文件扩展名自动选择解析器"""
    name = filename.lower()
    if name.endswith(".csv"):
        return parse_csv(file_bytes)
    return parse_excel(file_bytes)
